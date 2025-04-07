import openpyxl
import csv
import os

def extract_hyperlinks_from_column_g(file_path="./csv/teuchitlan.xlsx"):
    """
    Extract hyperlinks from column G of an Excel file.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        list: List of hyperlinks found in column G
    """
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    
    # Get the active worksheet (or first sheet)
    worksheet = workbook.active
    
    # Initialize list to store hyperlinks
    hyperlinks = []
    
    # Iterate through column G (column 7)
    for row in range(1, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=7)  # Column G is the 7th column
        
        # Check if cell has a hyperlink
        if cell.hyperlink:
            hyperlinks.append({
                'row': row,
                'display_text': cell.value,
                'hyperlink': cell.hyperlink.target
            })
    
    return hyperlinks

def extract_all_data_from_excel(file_path="./csv/teuchitlan.xlsx"):
    """
    Extract all data from an Excel file including header row.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        list: List of rows with all data from the Excel file
        list: Header row with column names
    """
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    
    # Get the active worksheet (or first sheet)
    worksheet = workbook.active
    
    # Extract header row (first row)
    header = [cell.value for cell in worksheet[1]]
    
    # Extract all data rows (including header)
    all_data = []
    for row_idx, row in enumerate(worksheet.rows, 1):
        row_data = [cell.value for cell in row]
        all_data.append({
            'row_num': row_idx,
            'data': row_data
        })
    
    # Identify significant columns (not empty)
    significant_columns = []
    for col_idx in range(len(header)):
        # Check if this column has any non-empty values
        has_data = False
        for row in all_data:
            if row_idx > 1 and col_idx < len(row['data']) and row['data'][col_idx] is not None:
                has_data = True
                break
        
        # If it's one of the first 7 columns or has data, keep it
        if col_idx < 7 or has_data:
            significant_columns.append(col_idx)
    
    # Filter the header and data to include only significant columns
    filtered_header = [header[i] for i in significant_columns]
    
    for row in all_data:
        row['data'] = [row['data'][i] if i < len(row['data']) else None for i in significant_columns]
    
    return all_data, filtered_header

def export_hyperlinks_to_csv(hyperlinks, output_file="./csv/hyperlinks3.csv"):
    """
    Export hyperlinks to a CSV file with an ID column.
    
    Args:
        hyperlinks (list): List of hyperlink dictionaries
        output_file (str): Path to the output CSV file
    """
    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    # Write to CSV file
    with open(output_file, 'w', newline='') as csvfile:
        fieldnames = ['id', 'row', 'display_text', 'hyperlink']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        
        writer.writeheader()
        for link in hyperlinks:
            # Add the id field with value '001' to each row
            writer.writerow({
                'id': '001',
                'row': link['row'],
                'display_text': link['display_text'],
                'hyperlink': link['hyperlink']
            })
    
    return output_file

def export_data_to_csv(all_data, header, hyperlinks, output_file="./csv/complete_data3.csv"):
    """
    Export all data to a CSV file with an incrementing ID column and hyperlink information.
    
    Args:
        all_data (list): List of all rows from Excel
        header (list): Header row with column names
        hyperlinks (list): List of hyperlink dictionaries
        output_file (str): Path to the output CSV file
    """
    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    # Create a dictionary for quick lookup of hyperlinks by row number
    hyperlink_dict = {link['row']: link for link in hyperlinks}
    
    # Prepare header row with additional columns
    csv_header = ['id'] + header + ['hyperlink']
    
    # Write to CSV file
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(csv_header)
        
        # Start ID counter (skip header row)
        id_counter = 1
        
        for row_data in all_data:
            row_num = row_data['row_num']
            data = row_data['data']
            
            # Skip header row (row 1)
            if row_num == 1:
                continue
                
            # Format ID as '001', '002', etc.
            formatted_id = f"{id_counter:03d}"
            
            # Add ID column
            row_to_write = [formatted_id] + data
            
            # Add hyperlink if available for this row
            if row_num in hyperlink_dict:
                row_to_write.append(hyperlink_dict[row_num]['hyperlink'])
            else:
                row_to_write.append('')
                
            writer.writerow(row_to_write)
            
            # Increment ID counter
            id_counter += 1
    
    return output_file

def main():
    # Extract hyperlinks from the Excel file
    hyperlinks = extract_hyperlinks_from_column_g()
    
    # Extract all data from the Excel file
    all_data, header = extract_all_data_from_excel()
    
    # Export complete data to CSV with hyperlinks and incremental IDs
    csv_file = export_data_to_csv(all_data, header, hyperlinks)
    
    print(f"Exported complete data with hyperlinks to {csv_file}")
    
    # Original hyperlinks CSV export kept for backward compatibility
    export_hyperlinks_to_csv(hyperlinks)

if __name__ == "__main__":
    main()
